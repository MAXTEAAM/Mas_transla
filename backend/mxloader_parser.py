"""
MXLoader Excel Parser & Exporter
=================================
Handles parsing MXLoader template Excel files from Maximo MAS 9.

MXLoader sheets typically have:
- Identifier columns (OBJECTNAME, ATTRIBUTENAME, MAXVALUE, VALUE, etc.)
- Language columns identified by ISO codes (EN, FR, DE, ES, AR, etc.)
  or full names (DESCRIPTION_EN, DESCRIPTION_FR, LONGDESCRIPTION_EN, etc.)
- Multiple sheets per workbook (one per object structure or domain)

This parser auto-detects the structure and extracts translatable content.
"""

import re
from copy import copy
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import get_column_letter


# Common Maximo language column patterns
LANG_CODES = {
    "EN", "FR", "DE", "ES", "IT", "PT", "NL", "AR", "ZH", "JA", "KO",
    "RU", "PL", "SV", "DA", "FI", "NO", "TR", "HE", "TH", "VI",
    "CS", "HU", "RO", "BG", "HR", "SK", "SL", "UK", "EL", "ID", "MS",
}

# Patterns that indicate a language column in MXLoader
LANG_PATTERNS = [
    re.compile(r"^([A-Z]{2})$"),                          # "EN", "FR"
    re.compile(r"^DESCRIPTION_([A-Z]{2})$", re.I),        # DESCRIPTION_EN
    re.compile(r"^LONGDESCRIPTION_([A-Z]{2})$", re.I),    # LONGDESCRIPTION_EN
    re.compile(r"^LDTEXT_([A-Z]{2})$", re.I),             # LDTEXT_EN
    re.compile(r"^VALUE_([A-Z]{2})$", re.I),              # VALUE_EN
    re.compile(r"^DESCRIPTION\.([A-Z]{2})$", re.I),       # DESCRIPTION.EN
    re.compile(r"^LABEL_([A-Z]{2})$", re.I),              # LABEL_EN
    re.compile(r"^([A-Z]{2})_DESCRIPTION$", re.I),        # EN_DESCRIPTION
    re.compile(r"^([A-Z]{2})_LONGDESCRIPTION$", re.I),    # EN_LONGDESCRIPTION
]

# Columns that are typically identifiers (not translatable)
ID_COLUMN_PATTERNS = {
    "OBJECTNAME", "ATTRIBUTENAME", "MAXVALUE", "VALUE", "DOMAINID",
    "SITEID", "ORGID", "INTERNAL", "DEFAULTS", "CLASSSTRUCTUREID",
    "CLASSIFICATIONID", "PARENT", "HIERARCHYPATH", "ABORESSION",
    "OBJECTSTRUCTURE", "ACTION", "LANGUAGECODE", "HASLD",
    "DESCRIPTION", "LONGDESCRIPTION",  # bare ones without lang suffix are source
}


class MXLoaderParser:
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(filepath)
        self._column_map: dict = {}  # sheet -> {col_idx: (col_name, lang_code|None)}

    def parse(self) -> dict:
        """
        Parse all sheets and return structured translation data.
        Returns:
            {
                "sheets": ["Sheet1", ...],
                "languages": ["EN", "FR", ...],
                "source_lang": "EN",
                "id_columns": ["OBJECTNAME", ...],
                "rows": [
                    {
                        "index": 0,
                        "sheet": "Sheet1",
                        "row_num": 2,
                        "identifiers": {"OBJECTNAME": "ASSET", ...},
                        "translations": {"EN": "Asset", "FR": "Actif", ...},
                        "field_type": "DESCRIPTION",  # or "LONGDESCRIPTION"
                        "auto_translated": {},
                    },
                    ...
                ]
            }
        """
        all_rows = []
        all_languages = set()
        all_id_columns = set()
        sheets = []

        for sheet_name in self.wb.sheetnames:
            ws = self.wb[sheet_name]
            if ws.max_row is None or ws.max_row < 2:
                continue

            # Read header row
            headers = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=1, column=col).value
                headers.append(str(val).strip() if val else "")

            # Classify columns
            lang_columns = {}  # col_idx -> (header, lang_code)
            id_columns = {}    # col_idx -> header
            desc_columns = {}  # col_idx -> header (bare DESCRIPTION/LONGDESCRIPTION)

            for idx, header in enumerate(headers):
                if not header:
                    continue

                # Check if it's a language column
                lang_code = self._detect_lang_code(header)
                if lang_code:
                    lang_columns[idx] = (header, lang_code)
                    all_languages.add(lang_code)
                elif header.upper() in ID_COLUMN_PATTERNS:
                    if header.upper() in ("DESCRIPTION", "LONGDESCRIPTION"):
                        desc_columns[idx] = header
                    else:
                        id_columns[idx] = header
                        all_id_columns.add(header)
                else:
                    # Treat as ID column by default
                    id_columns[idx] = header
                    all_id_columns.add(header)

            # If no language columns found but we have DESCRIPTION/LONGDESCRIPTION,
            # treat those as the source language (EN) content
            if not lang_columns and desc_columns:
                for idx, header in desc_columns.items():
                    lang_columns[idx] = (header, "EN")
                    all_languages.add("EN")

            if not lang_columns:
                continue  # Skip sheets with no translatable content

            sheets.append(sheet_name)
            self._column_map[sheet_name] = {
                "lang_columns": lang_columns,
                "id_columns": id_columns,
                "headers": headers,
            }

            # Parse data rows
            for row_num in range(2, ws.max_row + 1):
                identifiers = {}
                translations = {}
                has_content = False

                for col_idx, header in id_columns.items():
                    val = ws.cell(row=row_num, column=col_idx + 1).value
                    identifiers[header] = str(val) if val is not None else ""

                for col_idx, (header, lang_code) in lang_columns.items():
                    val = ws.cell(row=row_num, column=col_idx + 1).value
                    if val is not None and str(val).strip():
                        translations[lang_code] = str(val).strip()
                        has_content = True

                if has_content:
                    # Determine field type
                    field_type = "DESCRIPTION"
                    for _, (header, _) in lang_columns.items():
                        if "LONG" in header.upper():
                            field_type = "LONGDESCRIPTION"
                            break

                    all_rows.append({
                        "index": len(all_rows),
                        "sheet": sheet_name,
                        "row_num": row_num,
                        "identifiers": identifiers,
                        "translations": translations,
                        "field_type": field_type,
                        "auto_translated": {},
                    })

        # Determine source language (most populated one, usually EN)
        lang_counts = {}
        for row in all_rows:
            for lang in row["translations"]:
                lang_counts[lang] = lang_counts.get(lang, 0) + 1

        source_lang = max(lang_counts, key=lang_counts.get) if lang_counts else "EN"

        return {
            "sheets": sheets,
            "languages": sorted(all_languages),
            "source_lang": source_lang,
            "id_columns": sorted(all_id_columns),
            "rows": all_rows,
        }

    def _detect_lang_code(self, header: str) -> Optional[str]:
        """Try to extract a language code from a column header."""
        for pattern in LANG_PATTERNS:
            match = pattern.match(header)
            if match:
                code = match.group(1).upper()
                if code in LANG_CODES:
                    return code
        return None

    def export(self, rows: list, output_path: str):
        """
        Export updated translations back to MXLoader-compatible Excel.
        Preserves original formatting and structure, only updates translation cells.
        """
        # Copy the original workbook
        wb = openpyxl.load_workbook(self.filepath)

        for sheet_name, col_info in self._column_map.items():
            ws = wb[sheet_name]
            lang_columns = col_info["lang_columns"]

            # Build a lookup: (sheet, row_num) -> row data
            sheet_rows = {r["row_num"]: r for r in rows if r["sheet"] == sheet_name}

            for row_num, row_data in sheet_rows.items():
                for col_idx, (header, lang_code) in lang_columns.items():
                    new_val = row_data["translations"].get(lang_code)
                    if new_val:
                        ws.cell(row=row_num, column=col_idx + 1, value=new_val)

        wb.save(output_path)
        wb.close()
